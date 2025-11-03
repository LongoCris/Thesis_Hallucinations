import pandas as pd
from nltk.tokenize import sent_tokenize
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher


def build_tracciamento_catene(versioni_step):
    tracciamento = {}

    for i in range(len(versioni_step) - 1):
        step = f"step{i+1}"
        originale = sent_tokenize(versioni_step[i])
        modificato = sent_tokenize(versioni_step[i + 1])
        matcher = SequenceMatcher(None, originale, modificato)

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'replace':
                for orig, corr in zip(originale[i1:i2], modificato[j1:j2]):
                    frase_root = _find_root(tracciamento, orig)
                    if frase_root not in tracciamento:
                        tracciamento[frase_root] = []

                    tracciamento[frase_root].append({
                        "tipo": "modificata",
                        "frase": corr,
                        "step": step
                    })

            elif tag == 'delete':
                for orig in originale[i1:i2]:
                    frase_root = _find_root(tracciamento, orig)
                    if frase_root not in tracciamento:
                        tracciamento[frase_root] = []

                    tracciamento[frase_root].append({
                        "tipo": "rimossa",
                        "frase": "",
                        "step": step
                    })

            elif tag == 'insert':
                for corr in modificato[j1:j2]:
                    if corr not in tracciamento:
                        tracciamento[corr] = []
                    tracciamento[corr].append({
                        "tipo": "nuova",
                        "frase": corr,
                        "step": step
                    })



    return tracciamento


def _find_root(tracciamento, frase, visited=None):
    if visited is None:
        visited = set()
    if frase in visited:
        return frase  # stop ricorsione: ciclo rilevato
    visited.add(frase)

    for root, modifiche in tracciamento.items():
        for step in modifiche:
            if step["frase"] == frase:
                return _find_root(tracciamento, root, visited)

    return frase


def traccia_modifiche_excel(articolo_iniziale, articolo_finale, tracciamento_modifiche, path_output):
    frasi_finali = sent_tokenize(articolo_finale)
    dati = []

    for frase_originale, cambiamenti in tracciamento_modifiche.items():
        for cambio in cambiamenti:
            frase_corrente = cambio.get("frase", "")
            tipo = cambio.get("tipo", "")
            step = cambio.get("step", "")

            if frase_corrente and frase_corrente in articolo_finale:
                idx = articolo_finale.find(frase_corrente)
                start = max(0, idx - 30)
                end = idx + len(frase_corrente) + 30
                contesto = articolo_finale[start:end].replace("\n", " ").strip()
                presente = "si"
            else:
                contesto = ""
                presente = "no"

            dati.append({
                "frase radice": frase_originale,
                "frase": frase_corrente,
                "tipo": tipo,
                "step": step,
                "presente nel finale": presente,
                "contesto nel finale": contesto
            })


    df = pd.DataFrame(dati)

    # Scrittura Excel
    wb = Workbook()
    ws = wb.active
    ws.append(["frase radice", "frase", "tipo", "step", "presente nel finale", "contesto nel finale"])
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for row in df.itertuples(index=False):
        ws.append(list(row))
        idx = ws.max_row
        if row[4] == "si" and row[2]:
            ws.cell(row=idx, column=3).fill = green_fill

    ws.column_dimensions['F'].width = 60
    wb.save(path_output)
    print(f"✅ File Excel con catena modifiche salvato in: {path_output}")
