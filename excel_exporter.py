from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def esporta_excel(articolo_finale, articolo_zero, df_qa, correzioni_df, selezionati_info, domande_verificate, documenti, df_first, correzioni_first, selezionati_info_first, metodo_hallucination, tracciamento_fonti):
    from config import EXCEL_PATH

    nome_foglio = datetime.now().strftime("%Y-%m-%d %H.%M.%S")
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        wb.remove(wb.active)
    else:
        wb = load_workbook(EXCEL_PATH)

    ws = wb.create_sheet(title=nome_foglio)
    row_ptr = 4

    indice = [
        "Indice del contenuto:",
        "1. 🔍 Zero-check iniziale",
        "2. 🧪 First check concettuale",
        "3. 🗌 QA Dettagliati",
        "4. 📌 Articolo corretto finale",
        "5. ✅ Informazioni assenti selezionate manualmente",
        "6. 🔁 Iterazioni di correzione",
        "7. 🧠 Hallucinations identificate",
        "8. 📚 Tracciabilità delle fonti (Quarto check)"
    ]
    for i, voce in enumerate(indice):
        cell = ws.cell(row=i+1, column=7, value=voce)
        cell.font = Font(bold=i==0, color="000000")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    cell = ws.cell(row=1, column=1, value=f"Esecuzione del {nome_foglio}")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center")

    # Metodo hallucination
    row_ptr += 1
    ws.cell(row=row_ptr, column=1, value=f"Metodo di verifica hallucinations: {metodo_hallucination.upper()}").font = Font(bold=True, italic=True, color="333399")

    # Zero-check
    row_ptr += 3
    ws.cell(row=row_ptr, column=1, value="🔍 Zero-check: correzione iniziale dell'articolo").font = Font(bold=True, color="800000")
    for nome_file, contenuto in documenti.items():
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value=f"Fonte: {nome_file}").font = Font(italic=True, color="888888")
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value="📝 Articolo dopo correzione con questa fonte").font = Font(italic=True)
        for line in articolo_zero.splitlines():
            row_ptr += 1
            ws.cell(row=row_ptr, column=1, value=line)
        row_ptr += 2

    # First check
    row_ptr = ws.max_row + 2
    ws.cell(row=row_ptr, column=1, value="🧪 First check concettuale (QA concettuali)").font = Font(bold=True, color="004d00")
    headers = ["Domanda", "Risposta (Fonti)", "Risposta (Articolo)", "Giudizio", "Spiegazione"]
    row_ptr += 1
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=row_ptr, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    for idx, row in df_first.iterrows():
        row_ptr += 1
        ws.append([row[col] for col in headers])

    if selezionati_info_first:
        row_ptr += 2
        ws.cell(row=row_ptr, column=1, value="✅ Informazioni assenti selezionate manualmente (First Check)").font = Font(bold=True)
        for info in selezionati_info_first:
            row_ptr += 1
            ws.cell(row=row_ptr, column=1, value=info)

    # QA
    row_ptr += 2
    ws.cell(row=row_ptr, column=1, value="🗌 QA Dettagliati (frase per frase)").font = Font(bold=True, color="003366")
    headers = ["Domanda", "Risposta (Fonti)", "Risposta (Articolo)", "Giudizio", "Spiegazione"]
    row_ptr += 1
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=row_ptr, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCFF", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    for idx, row in df_qa.iterrows():
        row_ptr += 1
        ws.append([row[col] for col in headers])

    if selezionati_info:
        row_ptr += 2
        ws.cell(row=row_ptr, column=1, value="✅ Informazioni assenti selezionate manualmente (QA)").font = Font(bold=True)
        for info in selezionati_info:
            row_ptr += 1
            ws.cell(row=row_ptr, column=1, value=info)

    # Iterazioni
    row_ptr += 2
    ws.cell(row=row_ptr, column=1, value="🔁 Iterazioni di correzione").font = Font(bold=True)
    for i, row in enumerate(correzioni_df.itertuples(), 1):
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value=f"Iterazione {i}: Domanda → {row.Domanda}").font = Font(italic=True)
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value=f"Risposta corretta: {row._2}")

    # Articolo finale
    row_ptr += 2
    ws.cell(row=row_ptr, column=1, value="📌 Articolo corretto finale").font = Font(bold=True, color="000080")
    for line in articolo_finale.splitlines():
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value=line)

    # Hallucinations
    row_ptr += 2
    ws.cell(row=row_ptr, column=1, value="🧠 Hallucinations identificate").font = Font(bold=True)
    for domanda in domande_verificate:
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value=f"Domanda: {domanda}")

    # Quarto check
    row_ptr += 2
    ws.cell(row=row_ptr, column=1, value="📚 Tracciabilità delle fonti (Quarto check)").font = Font(bold=True, color="003366")
    headers = ["Frase", "Verifica"]
    row_ptr += 1
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=row_ptr, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EEE8AA", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    for item in tracciamento_fonti:
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value=item["Frase"])
        ws.cell(row=row_ptr, column=2, value=item["Verifica"])

    for i, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

    wb.save(EXCEL_PATH)
